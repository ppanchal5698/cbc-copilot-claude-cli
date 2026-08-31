"""Turning a catalog file into product records.

Pluggable and generic-first. Every vendor lays a price book out differently, so
these adapters read *structure* - a row with something code-shaped and something
money-shaped in it - rather than assuming column positions. A vendor-specific
adapter is worth writing only when the generic one demonstrably cannot cope, which
`ExtractionResult.validation_rate` is there to tell you.

Order matters: spreadsheets are already structured and are the cheapest, most
reliable rows in the corpus; a text PDF is next; OCR is the last resort because it
costs minutes per book and its output needs the most validation.
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Callable

import fitz  # PyMuPDF

from catalog_index.models import ExtractionResult, ProductRecord
from cbc_core import pdfrows

# Money: 1,234.56 or 1234.56, not a bare integer - a bare number in a price book is
# far more often a page number, a size or a quantity than a price.
_PRICE = re.compile(r"(?<![\d.])(\d{1,3}(?:,\d{3})*\.\d{2}|\d+\.\d{2})(?![\d])")

# Something code-shaped: has a digit, no spaces, not purely a number (a bare "2026"
# is a year, and "44" is a page).
_CODE = re.compile(r"^(?=[A-Za-z0-9])(?=.*\d)[A-Za-z0-9][A-Za-z0-9.\-/#]{2,29}$")

_UNITS = {"ea", "each", "pr", "pair", "set", "lf", "sf", "ft", "box", "cs", "case", "roll"}


def _price_in(text: str) -> float | None:
    found = _PRICE.findall(text or "")
    if not found:
        return None
    # The last money-shaped number on a row is the one in the price column; earlier
    # ones are sizes, dimensions or a list price the net column then discounts.
    return float(found[-1].replace(",", ""))


_STRIP = re.compile(r"[^A-Za-z0-9]")
_PHONE = re.compile(r"^\+?\d{3}[-. ]\d{3}[-. ]\d{4}$")

# A standalone number long enough to be a stock number rather than a page number, a
# year or a dimension. Hager's sheet leads each row with a six-digit item number.
MIN_NUMERIC_CODE = 5
# A numeric part number in the description position - Hager 3510, 3400, 3553 - is
# real and is what an architect specifies (CLAUDE.md, Phase 2).
MIN_DESCRIPTIVE_CODE = 3
# More words than this and the cell is a sentence, not a table cell. Terms and
# conditions pages are full of prices and phone numbers, and reading them as
# products floods search with boilerplate.
MAX_WORDS_IN_A_CELL = 8


def _is_code(candidate: str, *, minimum: int) -> bool:
    if not _CODE.match(candidate) or _PHONE.match(candidate):
        return False
    bare = _STRIP.sub("", candidate)
    if bare.isdigit():
        return len(bare) >= minimum
    return True


def _code_in(cells: list[str]) -> str | None:
    """The part number on this row, preferring the one an architect would specify.

    A whole cell is rarely the code. Hager prints rows as
    `010108 | BB1279 4-1/2" x 4-1/2" US26D | 4.42`, where the part number is the
    first token of the description and the leading number is Hager's own item
    number. Testing whole cells only found neither, so the largest vendor in the
    corpus indexed with no codes at all and exact SKU lookup for it always missed.

    Alphanumeric wins over purely numeric: architects specify "Hager 3400", not an
    internal stock number (see CLAUDE.md, Phase 2).
    """
    descriptive: str | None = None
    standalone: str | None = None

    for cell in cells:
        stripped = cell.strip()
        if not stripped:
            continue
        words = stripped.split()

        if len(words) == 1:
            # A cell that is only a code: Hager's leading item number, or a SKU.
            if _is_code(stripped, minimum=MIN_NUMERIC_CODE):
                if not _STRIP.sub("", stripped).isdigit():
                    return stripped
                standalone = standalone or stripped
        elif len(words) <= MAX_WORDS_IN_A_CELL:
            # A description: its first token is the manufacturer part number.
            head = words[0]
            if _is_code(head, minimum=MIN_DESCRIPTIVE_CODE):
                if not _STRIP.sub("", head).isdigit():
                    return head
                descriptive = descriptive or head

    # A numeric part in the description position beats the internal item number:
    # an architect writes "Hager 3510", not "035528".
    return descriptive or standalone


def _unit_in(cells: list[str]) -> str | None:
    for cell in cells:
        token = cell.strip().lower().rstrip(".")
        if token in _UNITS:
            return cell.strip().upper()
    return None


def _record_from_cells(cells: list[str], vendor: str, page: int) -> ProductRecord | None:
    """A row becomes a product when it is tabular *and* carries an identifier or price.

    The cell count is the load-bearing test. A price book's terms-and-conditions
    pages are full of prices ("a $50.00 flat fee") and things that look like codes
    (a phone number), and reading them as products buried the real parts under
    thousands of rows of boilerplate. A product lives in a table row with separate
    cells; a paragraph arrives as one.
    """
    filled = [c for c in cells if c.strip()]
    if len(filled) < 2:
        return None

    text = " | ".join(filled)
    code = _code_in(cells)
    price = _price_in(text)
    if code is None and price is None:
        return None  # a heading, a footnote, a column title

    # The description is the longest cell that is not the code or the price.
    parts = [
        c.strip() for c in cells
        if c.strip() and c.strip() != code and not _PRICE.fullmatch(c.strip())
    ]
    parts.sort(key=len, reverse=True)
    name = parts[0][:200] if parts else code
    detail = " ".join(parts[1:])[:500] or None

    return ProductRecord(
        vendor=vendor,
        page_number=page,
        product_code=code,
        name=name,
        description=detail,
        price=price,
        unit=_unit_in(cells),
        raw_text=text[:500],
    )


# ── adapters ───────────────────────────────────────────────────────────────


def extract_spreadsheet(path: Path, vendor: str) -> ExtractionResult:
    """Net sheets and cross-references. Already tabular, so the highest confidence."""
    from openpyxl import load_workbook

    result = ExtractionResult(extractor="spreadsheet")
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_number, sheet in enumerate(workbook.worksheets, start=1):
            result.pages_read += 1
            for row in sheet.iter_rows(values_only=True):
                cells = [str(value).strip() for value in row if value not in (None, "")]
                if not cells:
                    continue
                record = _record_from_cells(cells, vendor, sheet_number)
                if record:
                    result.records.append(record)
                else:
                    result.rejected["not_a_product_row"] = (
                        result.rejected.get("not_a_product_row", 0) + 1
                    )
    finally:
        workbook.close()
    return result


def _sniff_delimiter(sample: str) -> str:
    """The delimiter of a delimited file, defaulting to a comma when unsure."""
    try:
        return csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
    except csv.Error:
        return ","


def _read_delimited(path: Path) -> list[list[str]]:
    """Rows of a CSV/TSV, decoded tolerantly - latin-1 is the fallback that cannot fail."""
    for encoding in ("utf-8-sig", "latin-1"):
        try:
            text = path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        delimiter = "\t" if path.suffix.lower() == ".tsv" else _sniff_delimiter(text[:8192])
        return list(csv.reader(io.StringIO(text), delimiter=delimiter))
    return []


def extract_csv(path: Path, vendor: str) -> ExtractionResult:
    """Delimited net and cross-reference sheets. One table, read like a spreadsheet."""
    result = ExtractionResult(extractor="csv")
    result.pages_read = 1
    for row in _read_delimited(path):
        cells = [str(value).strip() for value in row if str(value).strip()]
        if not cells:
            continue
        record = _record_from_cells(cells, vendor, 1)
        if record:
            result.records.append(record)
        else:
            result.rejected["not_a_product_row"] = (
                result.rejected.get("not_a_product_row", 0) + 1
            )
    return result


def _is_description(cell: str) -> bool:
    """A multi-word cell whose first token is a part number - a product's own text."""
    words = cell.strip().split()
    return (
        2 <= len(words) <= MAX_WORDS_IN_A_CELL
        and _is_code(words[0], minimum=MIN_DESCRIPTIVE_CODE)
    )


def _split_columns(cells: list[str]) -> list[list[str]]:
    """Break one clustered row back into the products it actually contains.

    Hager prints two independent product tables side by side, so clustering by
    vertical position welds them into one row - which then pairs a part number from
    the left column with a price from the right. A wrong price against a real part
    number is the worst output this system can produce.

    The split is *not* on gap width, which was the obvious guess and is wrong:
    measured over seven sample pages the median gap inside a row is 94 pt while the
    gutter between the two columns is 53 pt, so the boundary is narrower than the
    spacing it would have to beat. What does mark it is the pattern - every product
    begins with its own description cell, so a second description starts a second
    product.
    """
    groups: list[list[str]] = []
    for cell in cells:
        if _is_description(cell) and any(_is_description(c) for c in (groups[-1] if groups else [])):
            groups.append([cell])
        elif groups:
            groups[-1].append(cell)
        else:
            groups.append([cell])
    return groups or [cells]


def extract_pdf_tables(path: Path, vendor: str, *, max_pages: int | None = None) -> ExtractionResult:
    """The generic PDF path: cluster positioned words into rows, then read the rows.

    Uses the same clustering and glyph repair as the bid-set take-off
    (`cbc_core/pdfrows.py`), so a price book whose fonts carry no ToUnicode map is
    recovered here exactly as it is there.
    """
    result = ExtractionResult(extractor="pdf_tables")
    doc = fitz.open(path)
    try:
        shift = pdfrows.detect_shift(doc, str(path))
        if shift:
            result.extractor = "pdf_tables+glyph_repair"
        limit = doc.page_count if max_pages is None else min(doc.page_count, max_pages)
        for index in range(limit):
            page = doc[index]
            result.pages_read += 1
            if not pdfrows.has_text_layer(page):
                result.rejected["no_text_layer"] = result.rejected.get("no_text_layer", 0) + 1
                continue
            for row in pdfrows.rows_from_words(page, shift=shift):
                for column in _split_columns(row["cells"]):
                    record = _record_from_cells(column, vendor, index + 1)
                    if record:
                        result.records.append(record)
    finally:
        doc.close()
    return result


def extract_pdf_ocr(path: Path, vendor: str, *, max_pages: int | None = None) -> ExtractionResult:
    """Scanned books. Minutes per file, so it never runs on a search path."""
    result = ExtractionResult(extractor="ocr")
    doc = fitz.open(path)
    try:
        limit = doc.page_count if max_pages is None else min(doc.page_count, max_pages)
        for index in range(limit):
            result.pages_read += 1
            text = pdfrows.ocr_page(doc[index])
            if text.startswith("[OCR "):
                result.rejected["ocr_unavailable"] = result.rejected.get("ocr_unavailable", 0) + 1
                continue
            for line in text.splitlines():
                cells = [c for c in re.split(r"\s{2,}", line.strip()) if c]
                record = _record_from_cells(cells, vendor, index + 1)
                if record:
                    result.records.append(record)
    finally:
        doc.close()
    return result


def needs_ocr(path: Path, sample_pages: int = 5) -> bool:
    """Classification: does this document have a usable text layer?"""
    if path.suffix.lower() != ".pdf":
        return False
    doc = fitz.open(path)
    try:
        pages = min(sample_pages, doc.page_count)
        if not pages:
            return False
        with_text = sum(1 for i in range(pages) if pdfrows.has_text_layer(doc[i]))
        return with_text == 0
    finally:
        doc.close()


SPREADSHEETS = {".xlsx", ".xlsm"}
# Delimited exports of the same net and cross-reference sheets. Tabular already,
# so they read row-for-row exactly like a spreadsheet.
CSV_FILES = {".csv", ".tsv"}
# Legacy .xls needs xlrd, which is a dependency for exactly one crossover sheet in
# this corpus. Reported as unsupported rather than failing the rebuild every time.
LEGACY_SPREADSHEETS = {".xls"}


def choose(path: Path) -> Callable[..., ExtractionResult]:
    """Which adapter reads this file."""
    if path.suffix.lower() in SPREADSHEETS:
        return extract_spreadsheet
    if path.suffix.lower() in CSV_FILES:
        return extract_csv
    if path.suffix.lower() in LEGACY_SPREADSHEETS:
        raise ValueError(
            f"{path.name} is the legacy .xls format, which openpyxl cannot read. "
            "Save it as .xlsx and it will index on the next rebuild."
        )
    if path.suffix.lower() == ".pdf":
        return extract_pdf_ocr if needs_ocr(path) else extract_pdf_tables
    raise ValueError(f"no extractor for {path.suffix!r} ({path.name})")
